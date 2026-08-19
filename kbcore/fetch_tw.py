"""台股官方端點與黃金保底的取數。

與 `fetch.py` 同一組具名例外，路由在 `fetch.py` 的 `route_of` 裡加。

## 單位陷阱

這一層存在的最大理由不是「發 HTTP 請求」，是**把單位換算收在一個地方**。
三大法人回的是元、融資金額是仟元、月營收是仟元——舊系統靠人記，而人會忘。
每一條路由的 `unit` 欄位就是它的宣告，寫進 raw 檔給下游看。

## 兩個 Chrome 時代的坑在這裡消失

- 「`navigate` 到 JSON 網址會被 Chrome 當檔案下載」——這一層不用 Chrome。
- 「跨子網域 `fetch` 被 CORS 擋」——伺服器端沒有同源政策。

**這是把採集從 Chrome 搬到 Actions 的真正收益**：那兩個坑不是被繞過，是不存在了。
"""
import csv
import io
import json
import urllib.error
import urllib.request
from typing import Dict

from .fetch import UA, TIMEOUT, ParseFailed, UnknownIdent, UpstreamError

ROUTES = {
    "TWSE:BFI82U": {
        "url": "https://www.twse.com.tw/rwd/zh/fund/BFI82U?dayDate={ymd}&type=day&response=json",
        "kind": "json", "unit": "元",
        "note": "三大法人買賣超。**單位是元**，下游要換算成億元。",
    },
    "TWSE:MI_MARGN": {
        "url": "https://www.twse.com.tw/rwd/zh/marginTrading/MI_MARGN?date={ymd}&selectType=MS&response=json",
        "kind": "json", "unit": "仟元",
        "note": "融資融券餘額。**融資金額單位是仟元。**",
    },
    "TPEX:INDEX": {
        "url": "https://www.tpex.org.tw/www/zh-tw/afterTrading/tradingIndex?date={y}/{m}/{d}&response=json",
        "kind": "json", "unit": "點",
        "note": "櫃買指數收盤。回傳當月逐日列，**取最後一列**。",
    },
    "TWSE:REV_L": {
        "url": "https://openapi.twse.com.tw/v1/opendata/t187ap05_L",
        "kind": "json", "unit": "仟元",
        "note": "上市公司月營收。**金額單位仟元。**",
    },
    "TWSE:REV_O": {
        "url": "https://mopsfin.twse.com.tw/opendata/t187ap05_O.csv",
        "kind": "csv", "unit": "仟元",
        "note": "上櫃公司月營收。**CSV，UTF-8 含 BOM。**",
    },
    "TWSE:CONF": {
        "url": "https://openapi.twse.com.tw/v1/opendata/t187ap04_L",
        "kind": "json", "unit": "—", "empty_ok": True,
        "note": ("法說會備援（主源是 MOPS 一覽表，需人工）。`符合條款` ＝ 第12款；"
                 "**`主旨 ` 這個欄位名結尾有一個半形空白**；週末回傳 0 筆——"
                 "那是「當日訊息」不是「行事曆」，0 筆是正常狀態不是故障。"),
    },
    "SPDR:GLD": {
        "url": ("https://api.spdrgoldshares.com/api/v1/historical-archive"
                "?product=gld&exchange=NYSE&lang=en"),
        "kind": "auto", "unit": "噸／美元",
        "note": ("GLD 歷史持倉序列。網址由使用者 2026-08-19 從官網的 Historical "
                 "Archive 下載連結取得——**我先前猜的 /assets/dynamic/ 路徑是錯的**，"
                 "它回了東西只是巧合（而且不是 UTF-8）。舊規格記載的 "
                 "`api.spdrgoldshares.com` 才是對的。"),
    },
    "SPDR:GLDM": {
        "url": ("https://api.spdrgoldshares.com/api/v1/historical-archive"
                "?product=gldm&exchange=NYSE&lang=en"),
        "kind": "auto", "unit": "噸／美元",
        "note": "同上，GLDM。",
    },
}


# 一天保留幾列。**截斷要看得見** —— raw 每天一個檔進 git，
# 完整歷史每天存一次會讓 repo 爆掉；但安靜的截斷比檔案大更危險。
XLSX_KEEP_LAST = 120


def parse_xlsx(raw: bytes, ident: str) -> Dict:
    """XLSX。**格式是實測出來的，不是從網頁文案推的。**

    2026-08-19：SPDR 的 historical-archive 端點回 content-type
    `application/vnd.openxmlformats-officedocument.spreadsheetml.sheet`、
    537KB、開頭 `50 4b 03 04`（ZIP）。這個結論來自 `sniff` 把 content-type 與
    開頭 bytes 印進失敗訊息——**猜格式的代價是症狀會指向錯的方向**。
    """
    try:
        import openpyxl
    except ImportError as e:
        raise ParseFailed(f"{ident} 是 XLSX，但這個環境沒有 openpyxl：{e}") from e
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        raise ParseFailed(f"{ident} XLSX 讀不開：{type(e).__name__}: {e}") from e

    # **「讀得開」不等於「有資料」。** 2026-08-19 第一版直接取 sheetnames[0]，
    # 拿到的是免責聲明頁：78 列全是 None，而狀態是 ok、必要項失敗是空的 ——
    # 每一個訊號都說成功，資料根本不存在。
    #
    # 這是同一個錯誤的第四次。前三次是「解得開」太寬鬆（latin-1、sniff、CSV 結構），
    # 這次是**「解析成功」太寬鬆**。共通的問法：我剛才那個 ok，是憑什麼判的？
    best, scored = None, []
    for name in wb.sheetnames:
        rows = [list(r) for r in wb[name].iter_rows(values_only=True)]
        n = sum(1 for r in rows for c in r if c is not None and str(c).strip())
        scored.append((name, len(rows), n))
        if best is None or n > best[2]:
            best = (name, rows, n)
    if best is None or best[2] == 0:
        raise ParseFailed(f"{ident} 每一個工作表都是空的：{scored}")

    name, rows, _ = best
    # 表頭是第一列**有三格以上非空**的列。免責聲明那種單格長文不算表頭。
    hi = next((i for i, r in enumerate(rows)
               if sum(1 for c in r if c is not None and str(c).strip()) >= 3), None)
    if hi is None:
        raise ParseFailed(f"{ident} 工作表 {name!r} 找不到表頭"
                          f"（沒有任何一列有三格以上非空）；各表：{scored}")
    header = [str(c).strip() if c is not None else "" for c in rows[hi]]
    body = [[c.isoformat() if hasattr(c, "isoformat") else c for c in r]
            for r in rows[hi + 1:]
            if any(c is not None and str(c).strip() for c in r)]
    if not body:
        raise ParseFailed(f"{ident} 工作表 {name!r} 表頭之後沒有任何資料列")

    kept = body[-XLSX_KEEP_LAST:]
    return {"sheet": name, "sheets": [x[0] for x in scored],
            "header_row": hi, "columns": header,
            "total_rows": len(body), "kept_last": len(kept),
            "dropped": len(body) - len(kept), "rows": kept}


def parse_json(raw: bytes, ident: str) -> Dict:
    try:
        return json.loads(raw.decode("utf-8-sig"))
    except (json.JSONDecodeError, UnicodeDecodeError) as e:
        raise ParseFailed(f"{ident} 回應不是合法 JSON：{e}") from e


# 依序試，第一個成功的就用。**用了哪一個要寫進輸出** ——
# 安靜的編碼退讓會讓「解對了」與「解成亂碼」長得一樣。
#
# **latin-1 刻意不在這裡。** 它能解開任何 byte 序列，所以放進來之後底下那個
# `raise ParseFailed` 就永遠不會執行——一個看起來像有處理、實際上是死碼的分支，
# 而且它的代價是把解不開的檔案變成亂碼而不是報錯。（2026-08-19 加編碼串聯時
# 第一版真的放了 latin-1，寫測試才發現那個分支打不到。）
#
# cp1252 仍然很寬鬆（只有五個未定義 byte 會擋），所以第二道防線是**把 encoding
# 寫進輸出並在摘要裡標出來**：退讓過就要看得見。
ENCODINGS = ("utf-8-sig", "cp1252")


def parse_csv(raw: bytes, ident: str) -> Dict:
    """UTF-8 BOM 由 `utf-8-sig` 吃掉；不是 UTF-8 的往下試 cp1252、latin-1。

    **欄位名一律原樣保留，不 strip。** `主旨 ` 結尾那個半形空白是真的，
    幫它清乾淨會讓下游照文件寫的欄位名反而取不到。

    編碼串聯是 2026-08-19 第一次實跑換來的：SPDR 的 GLD 歷史序列回了資料但
    `utf-8` 在 position 11 撞到 0xe2。**那個具名的 ParseFailed 是關鍵** ——
    它把「路徑錯了」與「路徑對但編碼不同」分開，兩者的處置完全不同。
    籠統的例外會讓這兩件事都顯示成「黃金失敗」。
    """
    text = enc_used = None
    for enc in ENCODINGS:
        try:
            text, enc_used = raw.decode(enc), enc
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ParseFailed(f"{ident} 用 {'、'.join(ENCODINGS)} 都解不開；"
                          f"開頭 32 bytes：{raw[:32].hex(' ')}")
    # **解得開不等於是 CSV。** cp1252 幾乎吃得下任何 byte，所以二進位檔（XLSX、
    # ZIP、PDF）會「成功」解成一行亂碼欄名然後回傳 —— 於是底下的失敗分支永遠打不到。
    #
    # 這是同一個錯誤的第三次（前兩次：latin-1 讓 ParseFailed 變死碼、sniff 的
    # raise 打不到）。共通形狀是：**我做的退讓太寬鬆，寬鬆到失敗分支不存在。**
    # 一條永遠不會執行的錯誤處理，比沒有錯誤處理更糟——它讓人以為處理過了。
    head = text[:2000]
    ctrl = sum(1 for c in head if ord(c) < 32 and c not in "\t\r\n")
    if "\x00" in head or ctrl > len(head) * 0.01:
        raise ParseFailed(
            f"{ident} 解得開但不是文字（前 2000 字裡有 {ctrl} 個控制字元）"
            f"—— 可能是二進位格式；開頭 32 bytes：{raw[:32].hex(' ')}")
    first = head.splitlines()[0] if head.splitlines() else ""
    if "," not in first and "\t" not in first:
        raise ParseFailed(f"{ident} 第一行沒有分隔符號，不是 CSV：{first[:60]!r}")

    r = csv.DictReader(io.StringIO(text))
    rows = list(r)
    if r.fieldnames is None:
        raise ParseFailed(f"{ident} 連表頭都沒有 —— 這不是空資料，是不是 CSV")
    # **零列不是錯誤。** 週末的法說會端點回 0 筆是正常狀態（「當日訊息」不是
    # 「行事曆」）。解析器報它看到什麼，「空算不算失敗」是呼叫端的政策 ——
    # 把政策烤進解析器，會讓正常狀態長得跟故障一模一樣。
    return {"encoding": enc_used, "columns": list(r.fieldnames), "rows": rows}


def get_tw(ident: str, ymd: str) -> Dict:
    """ymd 形如 20260819。回傳 {ident, url, unit, note, kind, data}。"""
    spec = ROUTES.get(ident)
    if spec is None:
        raise UnknownIdent(f"無法路由的代號：{ident!r}")

    url = spec["url"].format(ymd=ymd, y=ymd[:4], m=ymd[4:6], d=ymd[6:])
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    try:
        with urllib.request.urlopen(req, timeout=TIMEOUT) as r:
            raw = r.read()
            ctype = r.headers.get("Content-Type", "")
    except urllib.error.HTTPError as e:
        raise UpstreamError(f"{ident} 回 {e.code}") from e
    except Exception as e:
        raise UpstreamError(f"連不到 {ident}：{e}") from e

    if spec["kind"] == "auto":
        data, kind = sniff(raw, ident, ctype)
    else:
        kind = spec["kind"]
        data = parse_json(raw, ident) if kind == "json" else parse_csv(raw, ident)
    return {"ident": ident, "url": url, "unit": spec["unit"],
            "note": spec["note"], "kind": kind, "content_type": ctype, "data": data}


def sniff(raw: bytes, ident: str, ctype: str):
    """格式未知時，**讓程式報它拿到什麼**，不要由我猜。

    第一次接一個端點卻要先宣告它回 JSON 還是 CSV，就是在猜。猜錯的症狀是
    ParseFailed，而 ParseFailed 跟「路徑錯了」在「黃金卡出不來」那一層長得一樣。
    這裡把 content-type 與開頭的 bytes 一起寫進失敗訊息，下一輪就不用再猜。
    """
    if raw[:4] == b"PK\x03\x04":
        return parse_xlsx(raw, ident), "xlsx"
    for kind, fn in (("json", parse_json), ("csv", parse_csv)):
        try:
            return fn(raw, ident), kind
        except ParseFailed:
            continue
    raise ParseFailed(
        f"{ident} 既不是 JSON 也不是 CSV。content-type={ctype!r}、"
        f"{len(raw)} bytes、開頭 48 bytes：{raw[:48].hex(' ')}")
